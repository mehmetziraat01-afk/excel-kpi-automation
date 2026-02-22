"""DTM batch import service from Excel Load sheet."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.orm import Session

from yem_sistem.batch_items.models import BatchItem
from yem_sistem.imports.models import ImportJob, ImportStatus
from yem_sistem.materials.models import Material
from yem_sistem.production_batches.models import BatchStatus, ProductionBatch
from yem_sistem.stock_movements.models import MovementReason, MovementType, StockMovement
from yem_sistem.stock_movements.service import NegativeStockError, StockService


class DtmImportError(ValueError):
    """Base import validation error."""


@dataclass(slots=True)
class DtmImportSummary:
    rows_processed: int
    movements_created: int
    suspicious_batches_count: int


REQUIRED_COLUMNS = [
    "ID Batch",
    "Batch",
    "Date",
    "Start time",
    "End Time",
    "Feeder",
    "Recipe ID",
    "Recipe Name",
    "Ingredient Id",
    "Ingredient Name",
    "Target Weight",
    "Loaded",
    "Error (%)",
]


class DtmBatchImportService:
    SOURCE_NAME = "DTM_BATCH"

    def __init__(self, session: Session) -> None:
        self.session = session
        self.stock_service = StockService(session)

    def import_file(self, file_name: str, content: bytes, actor_role: str) -> DtmImportSummary:
        if actor_role.upper() != "ADMIN":
            raise PermissionError("Only ADMIN can import DTM batches.")
        if not file_name.lower().endswith((".xlsx", ".xls")):
            raise DtmImportError("Only .xls/.xlsx files are allowed")

        file_hash = hashlib.sha256(content).hexdigest()
        exists = self.session.execute(
            select(ImportJob.id).where(ImportJob.source_name == self.SOURCE_NAME, ImportJob.file_hash == file_hash)
        ).first()
        if exists is not None:
            raise DtmImportError("This file is already imported (hash duplicate).")

        import_job = ImportJob(source_name=self.SOURCE_NAME, file_name=file_name, file_hash=file_hash, status=ImportStatus.PENDING)
        self.session.add(import_job)
        self.session.flush()

        try:
            rows = self._parse_load_sheet(file_name=file_name, content=content)
            summary = self._persist_rows(rows)
            import_job.status = ImportStatus.SUCCESS
            import_job.message = (
                f"rows_processed={summary.rows_processed}, movements_created={summary.movements_created}, "
                f"suspicious_batches_count={summary.suspicious_batches_count}"
            )
            self.session.commit()
            return summary
        except Exception as exc:
            self.session.rollback()
            failed = ImportJob(
                source_name=self.SOURCE_NAME,
                file_name=file_name,
                file_hash=file_hash,
                status=ImportStatus.FAILED,
                message=str(exc),
            )
            self.session.add(failed)
            self.session.commit()
            raise

    def _parse_load_sheet(self, file_name: str, content: bytes) -> list[dict[str, object]]:
        if file_name.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ModuleNotFoundError as exc:
                raise DtmImportError("openpyxl is required for .xlsx import") from exc

            wb = load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb["Load"] if "Load" in wb.sheetnames else wb.active
            values = list(ws.iter_rows(values_only=True))
        else:
            try:
                import xlrd
            except ModuleNotFoundError as exc:
                raise DtmImportError("xlrd is required for .xls import") from exc

            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_name("Load") if "Load" in book.sheet_names() else book.sheet_by_index(0)
            values = [sheet.row_values(i) for i in range(sheet.nrows)]

        if not values:
            raise DtmImportError("Load sheet is empty")

        header = [str(c).strip() if c is not None else "" for c in values[0]]
        missing = [c for c in REQUIRED_COLUMNS if c not in header]
        if missing:
            raise DtmImportError(f"Missing required columns: {', '.join(missing)}")

        index = {name: header.index(name) for name in header}
        parsed: list[dict[str, object]] = []
        for row in values[1:]:
            if all((cell is None or str(cell).strip() == "") for cell in row):
                continue
            rec = {col: row[index[col]] if index[col] < len(row) else None for col in REQUIRED_COLUMNS}
            rec["Loaded DM KG (optional)"] = row[index["Loaded DM KG (optional)"]] if "Loaded DM KG (optional)" in index and index["Loaded DM KG (optional)"] < len(row) else None
            parsed.append(rec)
        return parsed

    def _persist_rows(self, rows: list[dict[str, object]]) -> DtmImportSummary:
        unknown_ingredients: set[str] = set()
        materials = list(self.session.scalars(select(Material)).all())
        by_code = {m.code.strip().upper(): m.id for m in materials}
        by_name = {m.name.strip().upper(): m.id for m in materials}

        def resolve_material_id(ingredient_id: object, ingredient_name: object) -> int | None:
            if ingredient_id is not None and str(ingredient_id).strip() != "":
                key = str(ingredient_id).strip().upper()
                if key in by_code:
                    return by_code[key]
            if ingredient_name is not None and str(ingredient_name).strip() != "":
                key = str(ingredient_name).strip().upper()
                if key in by_name:
                    return by_name[key]
            return None

        validated: list[dict[str, object]] = []
        for r in rows:
            loaded = self._to_decimal(r.get("Loaded"))
            if loaded is None or loaded < Decimal("0.000"):
                raise DtmImportError("Loaded value cannot be null or negative")

            material_id = resolve_material_id(r.get("Ingredient Id"), r.get("Ingredient Name"))
            if material_id is None:
                unknown_ingredients.add(f"{r.get('Ingredient Id')}/{r.get('Ingredient Name')}")
                continue

            validated.append({**r, "material_id": material_id, "loaded": loaded})

        if unknown_ingredients:
            raise DtmImportError(f"Unknown ingredients: {sorted(unknown_ingredients)}")

        batch_map: dict[tuple[str, date, time | None], ProductionBatch] = {}
        suspicious_batches: set[int] = set()
        movements_created = 0

        for r in validated:
            id_batch = str(r["ID Batch"]).strip()
            batch_date = self._to_date(r["Date"])
            start_time = self._to_time(r.get("Start time"))
            end_time = self._to_time(r.get("End Time"))
            key = (id_batch, batch_date, start_time)

            batch = batch_map.get(key)
            if batch is None:
                batch = ProductionBatch(
                    id_batch=id_batch,
                    batch_name=str(r.get("Batch") or "").strip(),
                    date=batch_date,
                    start_time=start_time,
                    end_time=end_time,
                    feeder=self._to_opt_str(r.get("Feeder")),
                    recipe_id=self._to_opt_str(r.get("Recipe ID")),
                    recipe_name=self._to_opt_str(r.get("Recipe Name")),
                    status=BatchStatus.OK,
                    suspicious_count_zero=0,
                )
                self.session.add(batch)
                self.session.flush()
                batch_map[key] = batch

            target_weight = self._to_decimal(r.get("Target Weight")) or Decimal("0.000")
            loaded = r["loaded"]
            error_percent = self._to_decimal(r.get("Error (%)"))

            item = BatchItem(
                production_batch_id=batch.id,
                material_id=int(r["material_id"]),
                id_batch=id_batch,
                start_time=start_time,
                target_weight=target_weight,
                loaded_weight=loaded,
                error_percent=error_percent,
                is_zero_loaded=(loaded == Decimal("0.000")),
            )
            self.session.add(item)

            if loaded == Decimal("0.000"):
                batch.status = BatchStatus.SUSPICIOUS
                batch.suspicious_count_zero += 1
                batch.suspicious_reason = "Contains zero loaded ingredient(s)."
                suspicious_batches.add(batch.id)
                continue

            movement_at = datetime.combine(batch_date, start_time or time(0, 0), tzinfo=timezone.utc)
            movement = StockMovement(
                material_id=int(r["material_id"]),
                movement_type=MovementType.OUT_PRODUCTION,
                reason=MovementReason.DTM_CONSUMPTION,
                quantity=loaded,
                movement_at=movement_at,
                reference_type="DTM_BATCH",
                reference_id=batch.id,
                note=f"id_batch={id_batch}",
            )
            try:
                self.stock_service.add_movement(movement)
            except NegativeStockError as exc:
                raise DtmImportError(str(exc)) from exc
            movements_created += 1

        return DtmImportSummary(
            rows_processed=len(validated),
            movements_created=movements_created,
            suspicious_batches_count=len(suspicious_batches),
        )

    @staticmethod
    def _to_opt_str(value: object) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        return s or None

    @staticmethod
    def _to_decimal(value: object) -> Decimal | None:
        if value is None:
            return None
        s = str(value).strip().replace(",", ".")
        if not s:
            return None
        return Decimal(s)

    @staticmethod
    def _to_date(value: object) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return datetime.fromisoformat(str(value)).date()

    @staticmethod
    def _to_time(value: object) -> time | None:
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, datetime):
            return value.time().replace(tzinfo=None)
        if isinstance(value, time):
            return value.replace(tzinfo=None)
        s = str(value).strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                pass
        return None
